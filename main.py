from evaluator import find_reasons
from colorama import init, Fore, Style
init(autoreset=True)
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
excel_path = "tests/mpin_logs.xlsx"

def run_checker(input_fn=input, output_fn=print):

    while True:
        output_fn("\n***Welcome to MPIN Checker***\n")
        output_fn("This tool is designed to check the strength of a given MPIN.")
        
        output_fn("\n***Please follow these rules when entering the dates***\n")
        output_fn("\n\t1.Enter the respective dates seperated by spaces (or any other character like '-' or '_' or '/')")
        output_fn("\n\t2.Each date should be in the format DD/MM/YYYY or MM/DD/YYYY or YYYY/MM/DD")
        output_fn("\n\t3.Please do not enter text or special characters instead of numbers in the inputs")

        output_fn("\nPlease provide the following details: \n")
        while True:
            mpin = input_fn("\nEnter the MPIN: ")
            if len(mpin) not in [4, 6]:  # to check if the mpin is 4 or 6 digits long, if not, ask for a new mpin
                output_fn("MPIN should be either 4 or 6 digits long. Please enter a new MPIN.")
                continue
            else:
                break

        dob_self = input_fn("Enter Date of Birth of Self: ")
        dob_spouse = input_fn("Enter Date of Birth of Spouse: ")
        anniversary = input_fn("Enter Anniversary Date: ")

        reasons = find_reasons(mpin, dob_self, dob_spouse, anniversary)
        strength = ""

        if "INVALID_DATE" in reasons:
            output_fn("Reasons Found: \n" + str(reasons))
            output_fn("Invalid Date Entered... \n Please Try Again...")

        elif len(reasons) == 0:
            output_fn(Fore.GREEN + "Your MPIN is considered STRONG")
            output_fn("Reasons Found: \n" + str(reasons))
            strength = "STRONG"

        elif len(reasons) <= 2:
            output_fn(Fore.YELLOW + "Your MPIN is considered WEAK ")
            output_fn("Reasons Found: \n" + str(reasons))
            strength = "WEAK"

        else:
            output_fn(Fore.RED + "Your MPIN is considered so WEAK ")
            output_fn("Reasons Found: \n" + str(reasons))
            strength = "VERY_WEAK"

        Style.RESET_ALL

        cont = input_fn("\nDo you want to check another MPIN? (yes/no): ")
        log_to_excel(mpin, dob_self, dob_spouse, anniversary, reasons, strength, cont)

        if cont == "n":
            break



from datetime import datetime
from openpyxl import load_workbook

excel_path = "tests/mpin_logs.xlsx"

def log_to_excel(mpin, dob_self, dob_spouse, anniversary, reasons, strength, cont):
    wb = load_workbook(excel_path)
    ws = wb["logs"]

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # Timestamp (first column)
        mpin,
        dob_self,
        dob_spouse,
        anniversary,
        ", ".join(reasons),
        strength,
        cont
    ])

    wb.save(excel_path)



if __name__ == "__main__":
    run_checker()
